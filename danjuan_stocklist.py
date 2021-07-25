#author : Joky
#date   : 2021/07/25
#weather: taifeng
from matplotlib.font_manager import FontProperties
from multiprocessing import Process, Queue
from win32com.client import Dispatch
from enum import IntEnum, Enum
import matplotlib.pyplot as plt
import requests
import openpyxl
import random
import json
import time
import sys
import os

#global var
width = 600
typeNum = timeNum = 0
is_quit = False
#const var
FUND_MODE = 6
OBSERVE_MODE = 8

USER_AGENTS = [
	"Mozilla/5.0 (Macintosh;Intel Mac OS X 10_7_3) AppleWebKit/535.20 (KHTML, like Gecko) Chrome/19.0.1036.7 Safari/535.20",
	"Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.1 (KHTML, like Gecho) Chrome/21.0.1180.71 Safari/537.1 LBBROWSER",
	"Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/535.11 (KHTML, like Gecho) Chrome/17.0.963.84 Safari/535.11 LBBROWSER",
	"MOzilla/4.0 (compatible; MSIE 7.0; Windows NT 6.1; WOW64; Trident/5.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; .NET4.0C; .NET4.0E)",
	"Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; QQDownload 732; .NET4.0C; .NET4.0E)",
	"Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Trident/4.0; SV1; QQDownload 732; .NET4.0C; .NET4.0E; 360SE)",
	"Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; QQDownload 732; .NET4.0C; .NET4.0E)",
	"Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.1 (KHTML, like Gecho) Chrome/21.0.1180.89 Safari/537.1",
	"Mozilla/5.0 (iPad; U; CPU OS 4_2_1 like Mac OS X; zh-cn) AppleWebKit/533.17.9 (KHTML, like Gecko) Version/5.0.2 Mobile/8C148 Safari/6533.18.5",
	"Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:2.0b13pre) Gecko/20110307 Firefox/4.0b13pre",
	"Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:16.0) Gecko/20100101 Firefox/16.0",
	"Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.11 (KHTML, like Gecho) Chrome/23.0.1271.64 Safari/537.11",
	"Mozilla/5.0 (X11; U; Linux x86_64; zh-CN; rv:1.9.2.10) Gecko/20100922 Ubuntu/10.10 (maverick) Firefox/3.6.10",
    "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36"
]

MY_JSON_CODES = '{\
    "codes" : ["005827", "110011", "003095", "001938", \
                "001811", "010213", "100032", "007994",\
                "001171", "161005", "001811", "001875", \
                "001216", "007130", "003984", "005609", \
                "161723", "160140"\
                ]}'

TEST_CODES = '{\
    "codes" : [\
        "005827", "007455", "004237", "002620", "001837", "000327",\
        "000326", "000940", "005598", "110011", "481010",\
        "100032", "008682", "090010", "161907",\
        "007994", "002906", "001351", "002311"\
    ]}'

font = FontProperties(fname = 'C:\\Windows\\Fonts\\msyh.ttc', size = 12)
class FundType(IntEnum):
    STOCK = 1
    MIX = 3
    BOND = 2
    QDII = 11
    SMART = 0
    DIY = 99

class ObserveTime(Enum):
    ONE_MONTH = '1m'
    THREE_MONTH = '3m'
    SIX_MONTH = '6m'
    ONE_YEAR = '1y'
    TWO_YEAR = '2y'
    THREE_YEAR = '3y'
    FIVE_YEAR = '5y'
    LAZY = '0'

#打开网站
def get_url_data(url, headers, params=None):
    try:
        response = requests.get(url=url, headers=headers, params=params)
        data = response.json()['data']
        return data
    except:
        print('url:%s, 获取数据失败。' % url)
        return None

#按名称排列、按基金经理排列
class Fund:
    def __init__(self, type, observe_time):
        self.type = type
        self.observe_type = observe_time
        self.headers = {
            'Cookie': 'device_id=web_SJedqKpesv; gr_user_id=1d04b762-0a3b-41c4-b8b8-004fba79b385; accesstoken=240010000d68ddef5d0a921adff3a9ca3b71caf188ee68894; u=161474641; uid=161474641; refreshtoken=2400100001755f51276cba8a0a5b949dff63d230d9bada5c7; acw_tc=2760778316125744776097384e603ef6fb50648ef4438b3511b678a989a199; xq_a_token=244708005bc946786b8ba870b872c392f4f1e35e; Hm_lvt_d8a99640d3ba3fdec41370651ce9b2ac=1612351930,1612574478; channel=1300100141; Hm_lpvt_d8a99640d3ba3fdec41370651ce9b2ac=1612574978; timestamp=1612576174413',
            'Referer': 'https://danjuanfunds.com',
            'User-Agent': random.choice(USER_AGENTS)
        }
        self.params = {
            'type' : self.type,
            'order_by' : self.observe_type,
            'size' : 20,
            'page' : 1
        }
        self.fd_codes = []
        self.fd_names = []
        #self.fd_date = []
        self.name_stock = {}
        self.manager_stock = {}
        self.stock_count = {}
        self.bond_count = {}

    def get_main(self):
        main_url = 'https://danjuanfunds.com/djapi/v3/filter/fund'
        set_page = 5
        for page in range(1, set_page + 1):
            self.params['page'] = page
            data = get_url_data(main_url, self.headers, self.params)
            if data == None:
                return
            if page == 1:
                if data['total_pages'] != None:
                    total_page = int(data['total_pages'])
                    set_page = min(set_page, total_page)
                else:
                    print('get total pages error.') 
            items = data['items']
            for item in items:
                self.fd_codes.append(item['fd_code'])
                self.fd_names.append(item['fd_name'])
                print('Get %s' % item['fd_name'])
            time.sleep(1)

    def get_stock_list(self):
        if not self.fd_codes:
            return False
        print('Wait some minutes is OK. \
            \nBecause it will visit many fund managers all around the world.')
        for i, code in enumerate(self.fd_codes):
            url = 'https://danjuanfunds.com/djapi/fund/detail/{}'.format(code)
            data = get_url_data(url, self.headers)
            if data == None:
                return
            try:
                fund_data = data['fund_position']
                stock_data = fund_data['stock_list']
                stock_list = [stock['name'] for stock in stock_data]  
            except Exception as e:
                print(e)
                continue
            try:
                bond_data = fund_data['bond_list']
                manager_data = data['manager_list']
                #self.fd_date.append(fund_data['enddate'])
                bond_list = [bond['name'] for bond in bond_data]
                manager_name = 'None'
                for mg in manager_data:
                    if manager_name == 'None':
                        manager_name = mg['name']
                    else:
                        manager_name = manager_name + ',' + mg['name']
            except Exception as e:
                print(e)
            self.count(stock_list, self.stock_count)
            self.merge_map_data(manager_name, stock_list, bond_list)
            if not self.fd_names:
                self.name_stock[code] = [stock_list, bond_list]
            else:
                self.name_stock[self.fd_names[i]] = [stock_list, bond_list]
            time.sleep(1) 
        print('Get data ok.')
        return True

    def count(self, mylist:list, mymap:map):
        for item in mylist:
            if item not in mymap.keys():
                mymap[item] = 1
            else:
                mymap[item] += 1

    def merge_map_data(self, mg_data, stocks, bonds):
        if mg_data in self.manager_stock.keys():
            norep_stock = [stock for stock in stocks if stock not in self.manager_stock[mg_data][0]]
            norep_bond = [bond for bond in bonds if bond not in self.manager_stock[mg_data][1]]
            self.manager_stock[mg_data][0].extend(norep_stock)
            self.manager_stock[mg_data][1].extend(norep_bond)
        else:
            self.manager_stock[mg_data] = [stocks, bonds]

    def get_name_stock(self):
        return self.name_stock

    #def get_enddate(self):
        #return self.fd_date

    def set_codes(self, codes):
        if codes != None:
            self.fd_codes.extend(codes)

    def get_manager_stock(self):
        return self.manager_stock

    def write_name_stock(self, file_name, sheet_name, head_names:tuple = ('Stock_Name', 'Stock')):
        self.write_to_excel(book_name=file_name, sheet_name=sheet_name, data_map=self.name_stock, head_names=head_names)

    def write_manager_stock(self, file_name, sheet_name, head_names:tuple = ('Manager', 'Stock')):
        self.write_to_excel(book_name=file_name, sheet_name=sheet_name, data_map=self.manager_stock, head_names=head_names)

    def write_to_excel(self, book_name, sheet_name, data_map, head_names:tuple, row=1, col=1):
        if not os.path.exists(book_name):
            wb = openpyxl.Workbook()
        else:
            wb = openpyxl.load_workbook(book_name)
        sheets = wb.worksheets
        if sheet_name not in sheets:
            wb.create_sheet(title=sheet_name,index=1)
        if 'Sheet' in sheets:
            ws = wb['Sheet']
            wb.remove(ws)
        ws = wb[sheet_name]
        #写入表头
        for i, head in enumerate(head_names):
            ws.cell(1, i+1).value = head
        #写入内容
        my_row = 2
        for name in data_map.keys():
            my_col = 1
            data = data_map[name]
            ws.cell(my_row, my_col).value = name
            for stock in data[0]:
                my_col += 1
                ws.cell(my_row, my_col).value = stock
            for bond in data[1]:
                my_col += 1
                ws.cell(my_row, my_col).value = bond
            my_row += 1
        wb.save(book_name)
        print('%s save ok.' % book_name)

    def get_image(self):
        plt.figure(figsize=(12, 7))
        demap = sorted(self.stock_count.items(), key=lambda x:x[1], reverse=True)
        height = [frq[1] for i, frq in enumerate(demap) if i < 50]
        labels = [frq[0] for i, frq in enumerate(demap) if i < 50]
        plt.bar([i for i in range(0, 50)], tuple(height), width=0.5, label="stock", color="#87CEFA")
        plt.xlabel('Stock_Name')
        plt.ylabel('Frequency')
        plt.title(get_sheet_name(self.type, self.observe_type))
        plt.xticks([i for i in range(0, 50)], tuple(labels), FontProperties = font, size = 8, rotation = 90)
        plt.show()

def open_file(file_path):
    if os.path.exists(file_path):
        try:
            os.startfile(file_path)
            print('%s Open Ok!' % file_path)
        except Exception as e:
            print('Can\'t open file. err:%s.' % str(e))

def get_path(file_name):
    path = os.getcwd()
    new_path = path + '\\' + 'Output\\' + file_name
    return new_path

def close_file(file_name):
    temp_file = '~$' + file_name
    path = get_path(temp_file)
    if os.path.exists(path):
        try:
            xl = Dispatch('Excel.Application')
            wb = xl.Workbooks.open(path)
            wb.Close(True)
        except Exception as e:
            print(e)

##########################初始化###########################
#初始化全局变量
def init_global():
    global width, typeNum, timeNum
    width = os.get_terminal_size().columns
    typeNum = timeNum = 0

def init_dir():
    if not os.path.exists('Output'):
        os.mkdir('Output')
    if not os.path.exists('json'):
        os.mkdir('json')
    with open('json//codes.json', 'w') as file:
        file.write(TEST_CODES)

#绘制菜单
def menu():
    global width
    #标题
    str = 'Welcome to Stock_List Hunter! '
    blankl = width // 2 - len(str) // 2 - 1
    blankr = width - len(str) - blankl - 2
    print('-' * width, end='')
    print('%s%s' % ('|', ' ' * blankl), end='')
    print(str, end='')
    print('%s%s' % (' ' * blankr, '|'), end='')
    print('-' * width)
    #选项
    sel_blank = width // 2 -4
    print('Stock Type: ')
    print('%s%s' % (' ' * sel_blank, '<1>.股票型'))
    print('%s%s' % (' ' * sel_blank, '<2>.混合型'))
    print('%s%s' % (' ' * sel_blank, '<3>.债券型'))
    print('%s%s' % (' ' * sel_blank, '<4>.QDII型'))
    print('%s%s' % (' ' * sel_blank, '<5>.SMART '))
    print('%s%s' % (' ' * sel_blank, '<6>.D I Y '))
    print('Observe Time: ')
    print('%s%s' % (' ' * sel_blank, '[1].一个月'))
    print('%s%s' % (' ' * sel_blank, '[2].三个月'))
    print('%s%s' % (' ' * sel_blank, '[3].六个月'))
    print('%s%s' % (' ' * sel_blank, '[4].一  年'))
    print('%s%s' % (' ' * sel_blank, '[5].两  年'))
    print('%s%s' % (' ' * sel_blank, '[6].三  年'))
    print('%s%s' % (' ' * sel_blank, '[7].五  年'))
    print('%s%s' % (' ' * sel_blank, '[8].TIUQ  '))
    #输入提示
    print('\ntake your choice (eg 1,2):')

#########################进程函数###########################
#获取选项
def get_choice(q, fn):
    sys.stdin = os.fdopen(fn)
    typeNum = timeNum = 0
    exp1 = exp2 = False
    while not (exp1 and exp2):
        try:
            typeNum, timeNum = eval(input())
        except:
            pass
        if q.empty():
            q.put(typeNum)
            q.put(timeNum)
        exp1 = typeNum > 0 and typeNum <= FUND_MODE
        exp2 = timeNum > 0 and timeNum <= OBSERVE_MODE
        time.sleep(0.8)

#########################其他函数###########################
#判断是否选择
def is_choice():
    global typeNum, timeNum, FUND_MODE, OBSERVE_MODE
    exp1 = typeNum > 0 and typeNum <= FUND_MODE
    exp2 = timeNum > 0 and timeNum <= OBSERVE_MODE
    if exp1 and exp2:
        return True
    else:
        return False

#获取控制台的宽度
def redraw_terminal(q):
    global width, typeNum, timeNum
    old_width = 0
    while not is_choice():
        width = os.get_terminal_size().columns
        if old_width != width:
            old_width = width
            os.system('cls' if os.name == 'nt' else 'clear')
            menu()
        if q.full():
            typeNum = q.get()
            timeNum = q.get()
        time.sleep(0.6)

def parse_selection():
    fund_type = {key : value for key, value in zip([i for i in range(1, FUND_MODE + 1)], FundType)}
    observe_time = {key : value for key, value in zip([i for i in range(1, OBSERVE_MODE + 1)], ObserveTime)}
    try:
        fundcode, observecode = fund_type[typeNum].value, observe_time[timeNum].value
    except:
        fundcode = -1
        observecode = 'err'
        #输出日志
    return fundcode, observecode

def get_fund_type(type):
    fund_types = []
    if type == 0:
        fund_types.extend([1, 2, 3, 11])
    else:
        fund_types.append(type)
    return fund_types

def get_sheet_name(type, observe_time):
    if type == FundType.STOCK.value:
        sheet_name = 'Stock'
    elif type == FundType.MIX.value:
        sheet_name = 'Mix'
    elif type == FundType.BOND.value:
        sheet_name = 'Bond'
    elif type == FundType.QDII.value:
        sheet_name = 'QDII'
    elif type == FundType.DIY.value:
        sheet_name = 'DIY'
    else:
        sheet_name = 'UnknowType'
    return sheet_name + str(observe_time)

def read_json(file_name='json//codes.json'):
    with open(file_name, 'r') as file:
        data = file.read()
        return json.loads(data)['codes']

def main():
    global is_quit
    fund_type, observe_type = parse_selection()
    if observe_type == ObserveTime.LAZY.value:
        print('!boj doog')
        is_quit = True
        return
    if fund_type == -1 or observe_type == 'err':
        return
    fund_types = get_fund_type(fund_type)
    file_name = ''
    for type in fund_types:
        my_fund = Fund(type, observe_type)
        if type != FundType.DIY.value:
            my_fund.get_main()
        else:
            my_fund.set_codes(read_json())
        bRes = my_fund.get_stock_list()
        if not bRes:
            continue
        if file_name == '':
            file_name = input('请输入欲保存的文件名(eg xxxx)：') + '.xlsx'
        file_path = get_path(file_name)
        sheet_name = get_sheet_name(type, observe_type)
        try:
            my_fund.write_name_stock(file_name=file_path, sheet_name=sheet_name+'ns')
            my_fund.write_manager_stock(file_name=file_path, sheet_name=sheet_name+'ms')
        except Exception as e:
            print(e)
            if os.path.exists(file_path):
                os.remove(file_path)
        my_fund.get_image()
    #打开文件
    close_file(file_name)
    open_file(file_path)
    tmp = input('Press any key to continue...')

if __name__ == '__main__':
    q = Queue(2)    #创建进程队列
    fn = sys.stdin.fileno() #输入重定向到子线程
    init_dir()
    while not is_quit:
        init_global()
        os.system('cls' if os.name == 'nt' else 'clear')
        pross = Process(target=get_choice, args=(q, fn))
        pross.daemon = True
        pross.start()
        redraw_terminal(q)
        pross.join()
        main()